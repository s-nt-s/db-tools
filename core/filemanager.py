import json
import logging
from os import W_OK, access, makedirs
from os.path import dirname, realpath
from pathlib import Path
from tempfile import gettempdir
import pickle
import csv as csvwriter
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from configparser import ConfigParser, MissingSectionHeaderError
from openpyxl.worksheet.worksheet import Worksheet

import pandas as pd

logger = logging.getLogger(__name__)


class FileManager:
    """
    Da funcionalidad de lectura (load) y escritura (dump) de ficheros
    sin importar el entorno de ejecución.
    Para ello transforma las rutas según necesario para poder escribir
    en directorios donde si haya permiso de escritura.
    """
    FM = None

    @staticmethod
    def get():
        """
        Devuelve una unica instancia de FileManager, con la configuracion por defecto
        """
        if FileManager.FM is None:
            FileManager.FM = FileManager()
        return FileManager.FM

    def __init__(self, root=None, scope: str = 'py.filemanager'):
        """
        Parameters
        ----------
        root: str | Path
            por defecto es la raiz del proyecto, es decir, el directorio ../.. al de este fichero
            se usa para interpretar que las rutas relativas son relativas a este directorio root
        scope: str
            nombre de subcarpeta a usar en tmp
        """
        if root is None:
            root = Path(dirname(realpath(__file__))).parent
        elif isinstance(root, str):
            root = Path(root)

        self.root = root
        self.temp = Path(gettempdir()) / scope

        for label, path in (
                ("raiz", self.root),
                ("temp", self.temp),
        ):
            wr = self.is_writeable(path)
            logger.info("Directorio %s %s [%s]",
                        label, path, "W_OK" if wr else "W_KO")

    @property
    def temp_root(self) -> Path:
        """
        Ruta donde se guardaran los ficheros que no se puedan crear en root
        """
        return self.temp / 'root'

    def is_writeable(self, path) -> bool:
        """
        Determina si se podra escribir un fichero en la ruta pasada por parametro
        """
        if isinstance(path, str):
            path = Path(path)
        while not (path.exists() or path.parent == path):
            path = path.parent
        return access(path, W_OK)

    def _resolve_path(self, file, wr: bool = False) -> Path:
        """
        Si es una ruta absoluta se devuelve tal cual
        Si es una ruta relativa y se requiere escribir en ella:
            * se devuelve bajo la ruta root si se puede escribir en ella
            * y si no se devuelve sobre la ruta temp
        Si es una ruta relativa y no se requiere escribir en ella:
            * se devuelve bajo la temp si el fichero existe
            * y si no se devuelve sobre la ruta root

        Parameters
        ----------
        file: str | Path
            Ruta a resolver
        wr: bool
            Indica si la ruta va a ser utilizada para escribir en ella o no
        """

        if isinstance(file, str):
            file = Path(file)

        if str(file).startswith("~"):
            file = file.expanduser()

        if file.is_absolute():
            return file

        temp_file = self.temp_root.joinpath(file)
        root_file = self.root.joinpath(file)
        if wr:
            if self.is_writeable(root_file):
                return root_file
            return temp_file

        if temp_file.exists():
            # Devolvemos el temporal porque
            # si el fichero temporal existe es que anteriormente se escribio en él,
            # si se escribio en el temporal es porque en el raiz no se pudo
            # por lo tanto el temporal es más actual que el que esta en el raiz
            return temp_file

        return root_file

    def resolve_path(self, file, *args, wr=False, **kvargs) -> Path:
        """
        Ver documentación _resolve_path
        """
        path = self._resolve_path(file, *args, **kvargs)
        if file != path and not (self.root.joinpath(file) == path):
            logger.info("[%s] %s -> %s", "RW"[int(wr)], file, path)
        return path

    def normalize_ext(self, ext) -> str:
        """
        Normaliza extensiones para identificar el tipo de fichero en base a la extension
        """
        ext = ext.lstrip(".")
        ext = ext.lower()
        return {
            "xlsx": "xls",
            "js": "json",
            "yml": "yaml",
            "sql": "txt"
        }.get(ext, ext)

    def exist(self, file, *args, **kvargs):
        return self.resolve_path(file).exists()

    def load(self, file, *args, **kvargs):
        """
        Lee un fichero en funcion de su extension
        Para que haya soporte para esa extension ha de existir una funcion load_extension
        """
        file = self.resolve_path(file)

        ext = self.normalize_ext(file.suffix)

        load_fl = getattr(self, "load_" + ext, None)
        if load_fl is None:
            raise Exception(
                "No existe metodo para leer ficheros {} [{}]".format(ext, file.name))

        return load_fl(file, *args, **kvargs)

    def dump(self, file, obj, *args, **kvargs):
        """
        Guarda un fichero en funcion de su extension
        Para que haya soporte para esa extension ha de existir una funcion dump_extension
        """
        file = self.resolve_path(file, wr=True)
        makedirs(file.parent, exist_ok=True)

        if len(args) == 0 and len(kvargs) == 0 and isinstance(obj, bytes):
            with open(file, "wb") as fl:
                fl.write(obj)
            return

        ext = self.normalize_ext(file.suffix)
        dump_fl = getattr(self, "dump_" + ext, None)
        if dump_fl is None:
            raise Exception(
                "No existe metodo para guardar ficheros {} [{}]".format(ext, file.name))

        dump_fl(file, obj, *args, **kvargs)

    def load_properties(self, file: Path, *args, **kvargs):
        try:
            config = ConfigParser()
            config.optionxform = str
            with open(file, "r") as f:
                config.read_file(f)
                return config
        except MissingSectionHeaderError:
            config = ConfigParser()
            config.optionxform = str
            with open(file, "r") as f:
                content = '[default]\n' + f.read()
                config.read_string(content)
                return config

    def load_json(self, file: Path, *args, **kvargs):
        with open(file, "r") as f:
            return json.load(f, *args, **kvargs)

    def dump_json(self, file: Path, obj, *args, indent=2, **kvargs):
        with open(file, "w") as f:
            json.dump(obj, f, *args, indent=indent, **kvargs)

    def load_csv(self, file: Path, *args, **kvargs):
        return pd.read_csv(file, *args, **kvargs)

    def dump_properties(self, file: Path, config: ConfigParser, *args, **kvargs):
        with open(file, "w") as f:
            config.write(f)
        
    def dump_csv(self, file: Path, obj, *args, **kvargs):
        if isinstance(obj, list):
            if len(obj) == 0:
                return
            keys = []
            for row in obj:
                for k in row.keys():
                    if k not in keys:
                        keys.append(k)
            with open(file, 'w', encoding='utf8', newline='') as f:
                dict_writer = csvwriter.DictWriter(f, keys)
                dict_writer.writeheader()
                dict_writer.writerows(obj)
            return
        obj.to_csv(file, *args, **kvargs)

    def dump_xls(self, file: Path, obj: pd.DataFrame, *args, prettify=False, **kvargs):
        max_rows = 200000 - 1
        if len(obj) > max_rows:
            count = 0
            name, ext = str(file).rsplit(".", 1)
            for i in range(0, len(obj), max_rows):
                count = count + 1
                fl = Path(f"{name}.{count:02d}.{ext}")
                self.dump_xls(fl, obj.iloc[i:(i + max_rows)], *args, prettify=prettify, **kvargs)
            return
    
        obj.to_excel(file, *args, **kvargs)
        if not prettify:
            return

        WB = load_workbook(file)
        for ws in WB.worksheets:
            if not(ws.max_row > 1 or ws.max_column > 1 or ws['A1'].value is not None):
                continue
            for i, col in enumerate(obj.columns):
                if kvargs.get('index') is not False:
                    i = i + 1
                cls = obj[col].dropna().drop_duplicates().values.tolist()
                if len(cls) > 0:
                    cls = list(map(lambda x: str(int(x)) if isinstance(x, (int, float)) else x, cls))
                cls.append(col)
                wdt = max(map(len, cls))
                l = get_column_letter(i + 1)
                w = max(wdt + 2, 6)
                ws.column_dimensions[l].width = w
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = get_column_letter(ws.max_column+1) + str(2)
        WB.save(file)

    def load_txt(self, file: Path, *args, **kvargs):
        with open(file, "r") as f:
            txt = f.read()
            if args or kvargs:
                txt = txt.format(*args, **kvargs)
            return txt

    def dump_txt(self, file: Path, txt, *args, **kvargs):
        if args or kvargs:
            txt = txt.format(*args, **kvargs)
        with open(file, "w") as f:
            f.write(txt)

    def load_pickle(self, file: Path, *args, **kvargs):
        with open(file, "rb") as f:
            return pickle.load(f)

    def dump_pickle(self, file: Path, obj, *args, **kvargs):
        with open(file, "wb") as f:
            pickle.dump(obj, f)


# Mejoras dinámicas en la documentación
FileManager.resolve_path.__doc__ = FileManager._resolve_path.__doc__
for mth in dir(FileManager):
    slp = mth.split("_", 1)
    if len(slp) == 2 and slp[0] in ("load", "dump"):
        key, ext = slp
        mth = getattr(FileManager, mth)
        if mth.__doc__ is None:
            if key == "load":
                mth.__doc__ = "Lee "
            else:
                mth.__doc__ = "Guarda "
            mth.__doc__ = mth.__doc__ + "un fichero de tipo " + ext

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    f = FileManager()
